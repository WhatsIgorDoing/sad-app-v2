import openpyxl
from openpyxl import Workbook

# Criar template com linha FIM conforme exemplo do usuário
wb = Workbook()
ws = wb.active

# Cabeçalho
ws["A1"] = "DOCUMENTO"
ws["B1"] = "REVISÃO"
ws["C1"] = "TÍTULO"
ws["D1"] = "ARQUIVO"
ws["E1"] = "FORMATO"
ws["F1"] = "DISCIPLINA"
ws["G1"] = "TIPO DE DOCUMENTO"
ws["H1"] = "PROPÓSITO"
ws["I1"] = "CAMINHO DATABOOK"

# Linha FIM (conforme exemplo fornecido)
ws["A2"] = "FIM"

wb.save("tests/fixtures/template_exemplo.xlsx")
print("Template atualizado com linha FIM")
