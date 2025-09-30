import sys

sys.path.insert(0, "src")

from pathlib import Path
import tempfile
import openpyxl

from sad_app_v2.core.domain import ManifestItem, DocumentFile, DocumentGroup
from sad_app_v2.infrastructure.file_system import SafeFileSystemManager
from sad_app_v2.infrastructure.template_filler import OpenpyxlTemplateFiller


def test_template_with_fim():
    print("=== TESTE TEMPLATE COM LINHA FIM ===")

    # 1. Verificar template base
    template_path = Path("tests/fixtures/template_exemplo.xlsx")
    print(f"Template base: {template_path}")

    wb_base = openpyxl.load_workbook(template_path)
    ws_base = wb_base.active
    print("\nTemplate base:")
    for i in range(1, ws_base.max_row + 1):
        row = [cell.value for cell in ws_base[i]]
        print(f"  Linha {i}: {row}")

    # 2. Criar dados de teste
    manifest_item = ManifestItem(
        document_code="RL-5290.00-22212-911-CZ6-019",
        revision="0",
        title="Inspeção de Recebimento-Emissão de RIR",
        metadata={
            "FORMATO": "A4",
            "DISCIPLINA": "ENGENHARIA DE PROJETO",
            "TIPO DE DOCUMENTO": "RL",
            "PROPÓSITO": "Para Construção",
            "CAMINHO DATABOOK": "DATA BOOK C&M",
        },
    )

    file1 = DocumentFile(
        Path("RL-5290.00-22212-911-CZ6-019_0.xlsx"),
        1000,
        associated_manifest_item=manifest_item,
    )
    file2 = DocumentFile(
        Path("RL-5290.00-22212-911-CZ6-019_A.pdf"),
        2000,
        associated_manifest_item=manifest_item,
    )

    group = DocumentGroup(
        document_code="RL-5290.00-22212-911-CZ6-019", files=[file1, file2]
    )

    # 3. Preencher template
    with tempfile.TemporaryDirectory() as temp_dir:
        output_path = Path(temp_dir) / "resultado_com_fim.xlsx"

        file_manager = SafeFileSystemManager()
        filler = OpenpyxlTemplateFiller(file_manager)

        filler.fill_and_save(template_path, output_path, [group])

        print(f"\n✅ Template preenchido: {output_path}")

        # 4. Verificar resultado final
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        print("\nTemplate preenchido:")
        for i in range(1, ws.max_row + 1):
            row = [cell.value for cell in ws[i]]
            if any(row):  # Só mostrar linhas não vazias
                print(f"  Linha {i}: {row}")

        # 5. Verificar se os dados estão entre cabeçalho e FIM
        print("\n=== VERIFICAÇÃO ESTRUTURAL ===")
        linha_1 = ws[1][0].value  # Primeira célula da linha 1

        # Encontrar linha FIM
        linha_fim = None
        for i in range(2, ws.max_row + 1):
            if ws[i][0].value == "FIM":
                linha_fim = i
                break

        print(f"Cabeçalho na linha 1: {linha_1}")
        print(f"Linha FIM encontrada em: {linha_fim}")

        if linha_fim:
            print(f"Dados entre linhas 2 e {linha_fim - 1}:")
            for i in range(2, linha_fim):
                row = [cell.value for cell in ws[i]]
                if any(row):
                    print(f"  Linha {i}: {row[0]} | {row[1]} | {row[2]} | {row[3]}")

            return True
        else:
            print("❌ Linha FIM não encontrada")
            return False


if __name__ == "__main__":
    success = test_template_with_fim()
    print(f"\n=== RESULTADO: {'SUCESSO' if success else 'FALHA'} ===")
