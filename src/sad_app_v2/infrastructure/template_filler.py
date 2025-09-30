# src/sad_app_v2/infrastructure/template_filler.py

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..core.domain import DocumentGroup
from ..core.interfaces import (
    IFileSystemManager,
    ITemplateFiller,
    TemplateFillError,
    TemplateNotFoundError,
)


def _get_filename_with_revision(original_filename: str, revision: str) -> str:
    """
    Constrói o nome do arquivo com a revisão adicionada antes da extensão.
    Verifica se a revisão já existe no nome para evitar duplicação.

    Exemplo:
    - "arquivo.pdf" + "A" -> "arquivo_A.pdf"
    - "arquivo_A.pdf" + "A" -> "arquivo_A.pdf" (não duplica)
    - "arquivo" + "B" -> "arquivo_B"
    """
    name_parts = original_filename.rsplit(".", 1)

    if len(name_parts) == 2:
        base_name, extension = name_parts
        # Verificar se já termina com _revisão
        if base_name.endswith(f"_{revision}"):
            # Já tem a revisão correta, retornar como está
            return original_filename
        else:
            # Adicionar a revisão
            return f"{base_name}_{revision}.{extension}"
    else:
        # Arquivo sem extensão
        if original_filename.endswith(f"_{revision}"):
            # Já tem a revisão correta, retornar como está
            return original_filename
        else:
            # Adicionar a revisão
            return f"{original_filename}_{revision}"


def _apply_header_formatting(sheet, header_row: int, num_columns: int):
    """Aplica formatação ao cabeçalho da planilha."""
    # Estilo do cabeçalho - fundo amarelo, texto em negrito
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=False
    )

    # Borda para todas as células
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Aplicar formatação ao cabeçalho - FORÇA aplicação célula por célula
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=header_row, column=col)
        # Limpar formatação existente primeiro
        cell.fill = PatternFill()  # Limpa
        cell.font = Font()  # Limpa
        cell.alignment = Alignment()  # Limpa
        cell.border = Border()  # Limpa

        # Aplicar nova formatação
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_data_formatting(sheet, start_row: int, end_row: int, num_columns: int):
    """Aplica formatação às linhas de dados."""
    # Alinhamento padrão para dados
    data_alignment = Alignment(horizontal="left", vertical="center")

    # Borda para todas as células
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Aplicar formatação aos dados
    for row in range(start_row, end_row + 1):
        for col in range(1, num_columns + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border


def _adjust_column_widths(sheet):
    """Ajusta as larguras das colunas baseado no conteúdo."""
    column_widths = {
        "A": 35,  # DOCUMENTO
        "B": 10,  # REVISÃO
        "C": 60,  # TÍTULO
        "D": 35,  # ARQUIVO
        "E": 10,  # FORMATO
        "F": 20,  # DISCIPLINA
        "G": 20,  # TIPO DE DOCUMENTO
        "H": 20,  # PROPÓSITO
        "I": 20,  # CAMINHO DATABOOK
    }

    # Forçar aplicação das larguras explicitamente
    for column, width in column_widths.items():
        dimension = sheet.column_dimensions[column]
        dimension.width = width
        # Garantir que a largura foi aplicada
        if dimension.width != width:
            dimension.width = width


class OpenpyxlTemplateFiller(ITemplateFiller):
    """Implementação que usa openpyxl para preencher templates Excel."""

    def __init__(self, file_manager: IFileSystemManager):
        self._file_manager = file_manager

    def fill_and_save(
        self, template_path: Path, output_path: Path, data: List[DocumentGroup]
    ) -> None:
        if not template_path.exists():
            raise TemplateNotFoundError(
                f"Arquivo template não encontrado: {template_path}"
            )

        # 1. Copia o template para o local de saída
        self._file_manager.copy_file(template_path, output_path)

        # 2. Abre a cópia e a preenche
        try:
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active

            # Encontrar onde inserir os dados (após cabeçalho, antes de "FIM")
            insert_row = 2  # Por padrão, inserir na linha 2 (após cabeçalho)

            # Procurar pela linha "FIM" para inserir antes dela
            for row_num in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_num, column=1).value
                if cell_value and str(cell_value).upper() == "FIM":
                    insert_row = row_num
                    break

            # Preparar todos os dados primeiro
            all_rows_data = []
            for group in data:
                # O ManifestItem é o mesmo para todos os arquivos em um grupo
                manifest_info = group.files[0].associated_manifest_item
                if not manifest_info:
                    continue  # Pula se por algum motivo não houver info

                # Cria uma linha para cada arquivo físico no grupo
                for file in group.files:
                    # Construir nome do arquivo com revisão (como ele ficará após ser movido)
                    revision = manifest_info.revision
                    filename_with_revision = _get_filename_with_revision(
                        file.path.name, revision
                    )

                    # Ordem correta dos campos conforme template:
                    row_data = [
                        manifest_info.document_code,  # DOCUMENTO
                        manifest_info.revision,  # REVISÃO
                        manifest_info.title,  # TÍTULO
                        filename_with_revision,  # ARQUIVO (nome final com revisão)
                        manifest_info.metadata.get("FORMATO", "A4"),  # FORMATO
                        manifest_info.metadata.get("DISCIPLINA", ""),  # DISCIPLINA
                        manifest_info.metadata.get(
                            "TIPO DE DOCUMENTO", ""
                        ),  # TIPO DE DOCUMENTO
                        manifest_info.metadata.get("PROPÓSITO", ""),  # PROPÓSITO
                        manifest_info.metadata.get(
                            "CAMINHO DATABOOK", ""
                        ),  # CAMINHO DATABOOK
                    ]
                    all_rows_data.append(row_data)

            # Inserir todas as linhas necessárias ANTES da linha "FIM"
            if all_rows_data:
                # Inserir todas as linhas necessárias antes da linha "FIM"
                for i in range(len(all_rows_data)):
                    sheet.insert_rows(insert_row)

                # Agora preencher os dados (FIM foi empurrada para baixo)
                for i, row_data in enumerate(all_rows_data):
                    target_row = insert_row + i
                    # Preencher a linha com os dados
                    for col_num, value in enumerate(row_data, 1):
                        sheet.cell(row=target_row, column=col_num, value=value)

            # Aplicar formatação à planilha
            self._apply_formatting(sheet, insert_row, len(all_rows_data))

            workbook.save(output_path)

        except Exception as e:
            raise TemplateFillError(f"Falha ao preencher o template {output_path}: {e}")

    def _apply_formatting(self, sheet, data_start_row: int, num_data_rows: int):
        """Aplica toda a formatação necessária à planilha."""
        num_columns = 9  # Total de colunas no template

        # 1. SEMPRE ajustar larguras das colunas primeiro
        _adjust_column_widths(sheet)

        # 2. Formatação do cabeçalho (linha 1) - FORÇA aplicação
        _apply_header_formatting(sheet, 1, num_columns)

        # 3. Formatação dos dados (se houver)
        if num_data_rows > 0:
            data_end_row = data_start_row + num_data_rows - 1
            _apply_data_formatting(sheet, data_start_row, data_end_row, num_columns)

        # 4. Formatação da linha "FIM" (se existir)
        fim_row = data_start_row + num_data_rows
        if sheet.cell(row=fim_row, column=1).value == "FIM":
            _apply_data_formatting(sheet, fim_row, fim_row, num_columns)

        # 5. Forçar aplicação das larguras novamente (garantia)
        _adjust_column_widths(sheet)
