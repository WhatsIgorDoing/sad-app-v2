import sys

sys.path.insert(0, "src")

from pathlib import Path
import tempfile
import openpyxl

from sad_app_v2.core.domain import ManifestItem, DocumentFile, DocumentGroup
from sad_app_v2.infrastructure.file_system import SafeFileSystemManager
from sad_app_v2.infrastructure.template_filler import OpenpyxlTemplateFiller


def test_template_formatting():
    print("=== TESTE DE FORMATAÇÃO DO TEMPLATE ===")

    # Criar dados de teste
    manifest_item = ManifestItem(
        document_code="CZ6_RNEST_U22_3.1.1.1_CVL_RIR_BSE-BA-02-022-DKF-0067",
        revision="A",
        title="Relatório de Inspeção de Recebimento de Estrutura de Concreto Existente",
        metadata={
            "FORMATO": "A4",
            "DISCIPLINA": "CIVIL",
            "TIPO DE DOCUMENTO": "RIR",
            "PROPÓSITO": "PARA CONSTRUÇÃO",
            "CAMINHO DATABOOK": "DATA BOOK C&M",
        },
    )

    file1 = DocumentFile(
        Path("CZ6_RNEST_U22_3.1.1.1_CVL_RIR_BSE-BA-02-022-DKF-0067.pdf"),
        1000,
        associated_manifest_item=manifest_item,
    )
    file2 = DocumentFile(
        Path("CZ6_RNEST_U22_3.1.1.1_CVL_RIR_BSE-BA-02-022-DKF-0067.xlsx"),
        2000,
        associated_manifest_item=manifest_item,
    )

    group = DocumentGroup(
        document_code="CZ6_RNEST_U22_3.1.1.1_CVL_RIR_BSE-BA-02-022-DKF-0067",
        files=[file1, file2],
    )

    # Preencher template com formatação
    with tempfile.TemporaryDirectory() as temp_dir:
        template_path = Path("tests/fixtures/template_exemplo.xlsx")
        output_path = Path(temp_dir) / "template_formatado.xlsx"

        file_manager = SafeFileSystemManager()
        filler = OpenpyxlTemplateFiller(file_manager)

        filler.fill_and_save(template_path, output_path, [group])

        print(f"✅ Template formatado criado: {output_path}")

        # Verificar formatação
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        print("\n=== VERIFICAÇÃO DE FORMATAÇÃO ===")

        # Verificar cabeçalho (linha 1)
        header_cell = ws["A1"]
        print(f"Cabeçalho A1:")
        print(f"  - Valor: {header_cell.value}")
        print(
            f"  - Fundo: {header_cell.fill.start_color.rgb if header_cell.fill.start_color else 'Sem cor'}"
        )
        print(f"  - Negrito: {header_cell.font.bold}")
        print(f"  - Largura coluna A: {ws.column_dimensions['A'].width}")

        # Verificar dados (linha 2)
        data_cell = ws["A2"]
        print(f"\nDados A2:")
        print(f"  - Valor: {data_cell.value}")
        print(f"  - Alinhamento: {data_cell.alignment.horizontal}")
        print(f"  - Borda esquerda: {data_cell.border.left.style}")

        # Verificar linha FIM
        fim_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "FIM":
                fim_row = row
                break

        if fim_row:
            fim_cell = ws[f"A{fim_row}"]
            print(f"\nLinha FIM (linha {fim_row}):")
            print(f"  - Valor: {fim_cell.value}")
            print(f"  - Borda: {fim_cell.border.left.style}")

        print(f"\n=== ESTRUTURA FINAL ===")
        for i in range(1, ws.max_row + 1):
            row_values = [ws.cell(row=i, column=j).value for j in range(1, 10)]
            if any(row_values):
                print(
                    f"Linha {i}: {row_values[0]} | {row_values[1]} | {row_values[2][:30] if row_values[2] else ''}"
                )

        return True


if __name__ == "__main__":
    success = test_template_formatting()
    print(f"\n=== RESULTADO: {'SUCESSO' if success else 'FALHA'} ===")
