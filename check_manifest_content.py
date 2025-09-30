import sys

sys.path.insert(0, "src")

from pathlib import Path

import openpyxl


def check_manifest_content():
    print("=== CONTEÚDO DO MANIFESTO ===")

    manifest_path = Path("tests/fixtures/manifesto_exemplo.xlsx")
    wb = openpyxl.load_workbook(manifest_path)
    ws = wb.active

    print("Cabeçalhos:")
    headers = [cell.value for cell in ws[1]]
    print(f"  {headers}")

    print("\nDados (primeiras 5 linhas):")
    for i in range(2, min(ws.max_row + 1, 7)):
        row = [cell.value for cell in ws[i]]
        print(f"  Linha {i}: {row}")


if __name__ == "__main__":
    check_manifest_content()
