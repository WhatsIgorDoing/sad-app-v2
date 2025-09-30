import openpyxl
from pathlib import Path

# Verificar template atual
template_path = Path("tests/fixtures/template_exemplo.xlsx")
print(f"Verificando template: {template_path}")

if template_path.exists():
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    print("\nConteúdo do template:")
    for i in range(1, ws.max_row + 1):
        row = [cell.value for cell in ws[i]]
        if any(row):
            print(f"  Linha {i}: {row}")
else:
    print("Template não encontrado!")
