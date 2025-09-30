import sys

sys.path.insert(0, "src")

import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from sad_app_v2.core.domain import DocumentFile, DocumentGroup, ManifestItem
from sad_app_v2.infrastructure.file_system import SafeFileSystemManager
from sad_app_v2.infrastructure.template_filler import OpenpyxlTemplateFiller


def debug_formatting():
    print("=== DEBUG - PROBLEMAS DE FORMATAÇÃO ===")

    # Criar dados simples
    manifest_item = ManifestItem(
        document_code="TESTE-001",
        revision="A",
        title="Documento de Teste",
        metadata={
            "FORMATO": "A4",
            "DISCIPLINA": "TESTE",
            "TIPO DE DOCUMENTO": "DOC",
            "PROPÓSITO": "TESTE",
            "CAMINHO DATABOOK": "TESTE",
        },
    )

    file1 = DocumentFile(
        Path("teste.pdf"), 1000, associated_manifest_item=manifest_item
    )
    group = DocumentGroup(document_code="TESTE-001", files=[file1])

    with tempfile.TemporaryDirectory() as temp_dir:
        template_path = Path("tests/fixtures/template_exemplo.xlsx")
        output_path = Path(temp_dir) / "debug_formatacao.xlsx"

        # Testar formatação original
        file_manager = SafeFileSystemManager()
        filler = OpenpyxlTemplateFiller(file_manager)
        filler.fill_and_save(template_path, output_path, [group])

        # Verificar resultado
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        print("\n=== VERIFICAÇÃO DETALHADA ===")

        # 1. Verificar cabeçalho
        header_cell = ws["A1"]
        print(f"CABEÇALHO A1:")
        print(f"  Valor: {header_cell.value}")
        print(
            f"  Fundo: {header_cell.fill.start_color.rgb if header_cell.fill.start_color else 'Sem cor'}"
        )
        print(f"  Negrito: {header_cell.font.bold}")
        print(f"  Alinhamento horizontal: {header_cell.alignment.horizontal}")
        print(f"  Alinhamento vertical: {header_cell.alignment.vertical}")

        # 2. Verificar larguras das colunas
        print(f"\nLARGURAS DAS COLUNAS:")
        for letter in ["A", "B", "C", "D", "E"]:
            width = ws.column_dimensions[letter].width
            print(f"  Coluna {letter}: {width}")

        # 3. Verificar se formatação está sendo aplicada
        print(f"\nTOTAL DE LINHAS: {ws.max_row}")
        print(f"TOTAL DE COLUNAS: {ws.max_column}")

        # 4. Aplicar formatação manual para comparação
        print(f"\n=== TESTE MANUAL DE FORMATAÇÃO ===")

        # Aplicar formatação manual no cabeçalho
        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 10):  # 9 colunas
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Ajustar larguras manualmente
        column_widths = {
            "A": 35,
            "B": 10,
            "C": 60,
            "D": 35,
            "E": 10,
            "F": 20,
            "G": 20,
            "H": 20,
            "I": 20,
        }
        for letter, width in column_widths.items():
            ws.column_dimensions[letter].width = width

        # Salvar versão corrigida
        corrected_path = Path(temp_dir) / "debug_formatacao_corrigida.xlsx"
        wb.save(corrected_path)

        print(f"✅ Versão original: {output_path}")
        print(f"✅ Versão corrigida: {corrected_path}")

        # Verificar novamente após correção manual
        wb2 = openpyxl.load_workbook(corrected_path)
        ws2 = wb2.active

        header_cell2 = ws2["A1"]
        print(f"\nAPÓS CORREÇÃO MANUAL:")
        print(f"  Alinhamento A1: {header_cell2.alignment.horizontal}")
        print(f"  Largura coluna A: {ws2.column_dimensions['A'].width}")
        print(f"  Largura coluna C: {ws2.column_dimensions['C'].width}")


if __name__ == "__main__":
    debug_formatting()
