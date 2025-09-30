import sys

sys.path.insert(0, "src")

from pathlib import Path

import openpyxl


def check_base_template():
    print("=== VERIFICAÇÃO DO TEMPLATE BASE ===")

    template_path = Path("tests/fixtures/template_exemplo.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    print(f"Template: {template_path}")

    # Verificar cabeçalho existente
    header_cell = ws["A1"]
    print(f"\nCABEÇALHO A1 NO TEMPLATE BASE:")
    print(f"  Valor: {header_cell.value}")
    print(
        f"  Fundo: {header_cell.fill.start_color.rgb if header_cell.fill.start_color else 'Sem cor'}"
    )
    print(f"  Negrito: {header_cell.font.bold}")
    print(f"  Alinhamento: {header_cell.alignment.horizontal}")

    # Verificar larguras existentes
    print(f"\nLARGURAS EXISTENTES:")
    for letter in ["A", "B", "C", "D", "E"]:
        width = ws.column_dimensions[letter].width
        print(f"  Coluna {letter}: {width}")

    print(f"\nESTRUTURA:")
    for i in range(1, min(ws.max_row + 1, 5)):
        values = [ws.cell(row=i, column=j).value for j in range(1, 10)]
        print(f"  Linha {i}: {values}")


if __name__ == "__main__":
    check_base_template()
