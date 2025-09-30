import sys

sys.path.insert(0, "src")

from pathlib import Path
import tempfile
import openpyxl

from sad_app_v2.core.use_cases.validate_batch import ValidateBatchUseCase
from sad_app_v2.core.use_cases.organize_lots import OrganizeAndGenerateLotsUseCase
from sad_app_v2.infrastructure.excel_reader import ExcelManifestRepository
from sad_app_v2.infrastructure.file_system import (
    FileSystemFileRepository,
    SafeFileSystemManager,
)
from sad_app_v2.infrastructure.template_filler import OpenpyxlTemplateFiller
from sad_app_v2.infrastructure.services import GreedyLotBalancerService


def test_full_organization():
    print("=== TESTE COMPLETO DE ORGANIZAÇÃO ===")

    # Preparar arquivos de teste
    manifest_path = Path("tests/fixtures/manifesto_exemplo.xlsx")
    template_path = Path("tests/fixtures/template_exemplo.xlsx")
    source_dir = Path("tests/fixtures")

    print(f"Manifesto: {manifest_path}")
    print(f"Template: {template_path}")
    print(f"Arquivos fonte: {source_dir}")
    print()

    # Repositórios e serviços
    manifest_repo = ExcelManifestRepository()
    file_repo = FileSystemFileRepository()
    file_manager = SafeFileSystemManager()
    template_filler = OpenpyxlTemplateFiller(file_manager)
    balancer = GreedyLotBalancerService()

    # Casos de uso
    validate_use_case = ValidateBatchUseCase(manifest_repo, file_repo)
    organize_use_case = OrganizeAndGenerateLotsUseCase(
        balancer, file_manager, template_filler
    )

    try:
        # 1. Validar lote
        print("1. Validando arquivos...")
        validated, unrecognized = validate_use_case.execute(manifest_path, source_dir)
        print(f"   Validados: {len(validated)}")

        if len(validated) == 0:
            print("   ⚠️ Nenhum arquivo validado")
            return False

        # 2. Organizar em lotes
        print("\n2. Organizando em lotes...")
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir) / "organizacao"
            output_dir.mkdir()

            # Copiar arquivos para área temporária primeiro (simular arquivos reais)
            temp_source = Path(temp_dir) / "source"
            temp_source.mkdir()

            for doc in validated[:2]:  # Usar apenas 2 arquivos para teste
                temp_file = temp_source / doc.path.name
                temp_file.write_text(f"Conteúdo simulado de {doc.path.name}")
                doc.path = temp_file  # Atualizar caminho para o arquivo temporário

            result = organize_use_case.execute(
                validated_files=validated[:2],
                output_directory=output_dir,
                master_template_path=template_path,
                max_docs_per_lot=10,
                lot_name_pattern="LOTE_0001",
                start_sequence_number=1,
            )

            print(f"   Resultado: {result.success}")
            print(f"   Lotes criados: {result.lots_created}")
            print(f"   Arquivos movidos: {result.files_moved}")
            print(f"   Mensagem: {result.message}")

            # 3. Verificar estrutura criada
            print("\n3. Verificando estrutura criada...")
            lote_dir = output_dir / "LOTE_0001"
            template_file = lote_dir / "LOTE_0001.xlsx"

            print(f"   Diretório do lote: {lote_dir}")
            print(f"   Existe: {lote_dir.exists()}")

            if lote_dir.exists():
                arquivos = list(lote_dir.iterdir())
                print(f"   Arquivos no lote: {[f.name for f in arquivos]}")

                print(f"\n   Template: {template_file}")
                print(f"   Template existe: {template_file.exists()}")

                # 4. Verificar conteúdo do template
                if template_file.exists():
                    print("\n4. Verificando conteúdo do template...")
                    wb = openpyxl.load_workbook(template_file)
                    ws = wb.active

                    print("   Cabeçalho:")
                    headers = [cell.value for cell in ws[1]]
                    print(f"   {headers}")

                    print("\n   Dados nas linhas:")
                    for row_num in range(
                        1, min(ws.max_row + 1, 5)
                    ):  # Primeiras 5 linhas
                        row_data = [cell.value for cell in ws[row_num]]
                        print(f"   Linha {row_num}: {row_data}")

                    # Verificar se há dados além do cabeçalho
                    has_data = False
                    for row_num in range(2, ws.max_row + 1):
                        row_data = [cell.value for cell in ws[row_num]]
                        if any(row_data):
                            has_data = True
                            break

                    if has_data:
                        print("   ✅ Template contém dados!")
                        return True
                    else:
                        print("   ❌ Template NÃO contém dados - apenas cabeçalho")
                        return False
                else:
                    print("   ❌ Template não foi criado")
                    return False
            else:
                print("   ❌ Diretório do lote não foi criado")
                return False

    except Exception as e:
        print(f"❌ Erro durante o teste: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_full_organization()
    print(f"\n=== RESULTADO: {'SUCESSO' if success else 'FALHA'} ===")
