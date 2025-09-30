import sys

sys.path.insert(0, "src")

import tempfile
from pathlib import Path

import openpyxl

from sad_app_v2.core.use_cases.organize_lots import OrganizeAndGenerateLotsUseCase
from sad_app_v2.core.use_cases.validate_batch import ValidateBatchUseCase
from sad_app_v2.infrastructure.excel_reader import ExcelManifestRepository
from sad_app_v2.infrastructure.file_system import (
    FileSystemFileRepository,
    SafeFileSystemManager,
)
from sad_app_v2.infrastructure.services import GreedyLotBalancerService
from sad_app_v2.infrastructure.template_filler import OpenpyxlTemplateFiller


def test_final_formatting():
    print("=== TESTE FINAL - FORMATAÇÃO COMPLETA ===")

    with tempfile.TemporaryDirectory() as temp_dir:
        output_dir = Path(temp_dir) / "organizacao"

        # Executar organização completa
        manifest_repo = ExcelManifestRepository()
        file_repo = FileSystemFileRepository()

        # Validação
        validate_uc = ValidateBatchUseCase(manifest_repo, file_repo)
        validated_files, _ = validate_uc.execute(
            Path("tests/fixtures/manifesto_exemplo.xlsx"), Path("tests/fixtures")
        )

        # Organização
        balancer = GreedyLotBalancerService()
        file_manager = SafeFileSystemManager()
        template_filler = OpenpyxlTemplateFiller(file_manager)

        organize_uc = OrganizeAndGenerateLotsUseCase(
            balancer, file_manager, template_filler
        )
        result = organize_uc.execute(
            validated_files=validated_files,
            output_directory=output_dir,
            master_template_path=Path("tests/fixtures/template_exemplo.xlsx"),
            max_docs_per_lot=10,
            start_sequence_number=1,
            lot_name_pattern="LOTE_XXXX",
        )

        if result.success:
            # Verificar arquivo gerado
            lote_path = output_dir / "LOTE_0001" / "LOTE_0001.xlsx"
            if lote_path.exists():
                wb = openpyxl.load_workbook(lote_path)
                ws = wb.active

                print(f"📊 Arquivo gerado: {lote_path}")

                print("\n=== VERIFICAÇÃO FINAL DE FORMATAÇÃO ===")

                # 1. Verificar cabeçalho
                header_cells = []
                for col in range(1, 10):
                    cell = ws.cell(row=1, column=col)
                    header_cells.append(
                        {
                            "valor": cell.value,
                            "fundo": cell.fill.start_color.rgb
                            if cell.fill.start_color
                            else "Sem cor",
                            "negrito": cell.font.bold,
                            "alinhamento_h": cell.alignment.horizontal,
                            "alinhamento_v": cell.alignment.vertical,
                        }
                    )

                print("🎨 CABEÇALHO (linha 1):")
                for i, cell_info in enumerate(
                    header_cells[:5], 1
                ):  # Mostrar só 5 primeiras
                    print(f"  Coluna {i}: {cell_info['valor']}")
                    print(f"    - Fundo: {cell_info['fundo']}")
                    print(f"    - Negrito: {cell_info['negrito']}")
                    print(f"    - Alinhamento: {cell_info['alinhamento_h']}")

                # 2. Verificar larguras das colunas
                print("\n📏 LARGURAS DAS COLUNAS:")
                expected_widths = [35, 10, 60, 35, 10, 20, 20, 20, 20]
                columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

                all_widths_correct = True
                for i, (col, expected) in enumerate(zip(columns, expected_widths)):
                    actual = ws.column_dimensions[col].width
                    status = "✅" if actual == expected else "❌"
                    print(f"  {col}: {actual} (esperado: {expected}) {status}")
                    if actual != expected:
                        all_widths_correct = False

                # 3. Verificar dados
                print("\n📄 DADOS:")
                data_row = 2
                data_cell = ws.cell(row=data_row, column=1)
                print(f"  Linha {data_row}: {data_cell.value}")
                print(f"    - Alinhamento: {data_cell.alignment.horizontal}")
                print(f"    - Borda: {data_cell.border.left.style}")

                # 4. Resumo final
                header_formatted = all(
                    cell["fundo"] == "00FFFF00"
                    and cell["negrito"]
                    and cell["alinhamento_h"] == "center"
                    for cell in header_cells
                )

                print(f"\n🔍 RESUMO FINAL:")
                print(f"  ✅ Cabeçalho formatado: {header_formatted}")
                print(f"  ✅ Larguras corretas: {all_widths_correct}")
                print(
                    f"  ✅ Arquivos com revisão: {any('_A.' in str(f) for f in lote_path.parent.iterdir())}"
                )
                print(
                    f"  ✅ Linha FIM preservada: {ws.cell(row=ws.max_row, column=1).value == 'FIM'}"
                )

                return header_formatted and all_widths_correct
            else:
                print("❌ Arquivo do lote não encontrado")
                return False
        else:
            print(f"❌ Organização falhou: {result.message}")
            return False


if __name__ == "__main__":
    success = test_final_formatting()
    print(f"\n🎉 RESULTADO FINAL: {'SUCESSO COMPLETO' if success else 'FALHA'} 🎉")
